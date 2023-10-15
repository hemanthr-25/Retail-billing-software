from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk,messagebox
import math
from datetime import date

# variables
count = 0
dic_prod={}
lst_key_prod=list(dic_prod.keys())

# workbook variables
wb1 = load_workbook("data.xlsx")
wb2 = load_workbook("inventory.xlsx")
ws1_wb2 = wb2["activeinventory"]

# inputing current date
c_d = date.today()

def popup(a):
    if a==0:
        messagebox.showerror("Inventory","Not enough stock")
    elif a==1:
        messagebox.showwarning("Inventory","Invalid Entry")

def reset():
    ivt_item_combo.delete(0, END)
    ivt_item_price.delete(0, END)
    ivt_item_qty.delete(0, END)
    ivt_item_qty.insert(0, 0)
    ivt_item_price.insert(0, 0)

def updatedict():
    global lst_key_prod
    row_max_ws1_wb2 = int(ws1_wb2.max_row)
    for p in range(3, row_max_ws1_wb2+1):
        entry2 = ws1_wb2.cell(row=p, column=1).value
        price = ws1_wb2.cell(row=p, column=3).value
        dic_prod[entry2]=int(price)
        lst_key_prod=list(dic_prod.keys())

def clc1():
    total_label.config(text="Please add items")
    for record in data_tree.get_children():
        data_tree.delete(record)

def ivtrefreshbtn():
    for record in ivt_data_tree.get_children():
        ivt_data_tree.delete(record)
    count2 = 1
    row_max_ws1_wb2 = int(ws1_wb2.max_row)
    for i in range(3, row_max_ws1_wb2+1):
        itemname_ws1_wb2 = ws1_wb2.cell(row=i, column=1).value
        itemqty_ws1_wb2 = ws1_wb2.cell(row=i, column=2).value
        itemprice_ws1_wb2 = ws1_wb2.cell(row=i, column=3).value
        ivt_data_tree.insert(parent="",index="end",iid=count2,text="",
        values=(itemname_ws1_wb2, itemqty_ws1_wb2,itemprice_ws1_wb2))
        count2 += 1

def homebtn():
    notebook.hide(1)
    notebook.select(0)

def inventorybtn():
    notebook.hide(0)
    notebook.select(1)
    ivtrefreshbtn()

def sheet_exists(td):
    wb1_sheetnames = wb1.sheetnames
    for names in wb1_sheetnames:
        if names == str(td):
            return True
        else:
            continue

def updateinv():
    global lst_key_prod
    global dic_prod
    entry2 = ivt_item_combo.get()
    ivtqty = ivt_item_qty.get()
    price=ivt_item_price.get()
    row_max_ws1_wb2 = int(ws1_wb2.max_row)
    for p in range(3, row_max_ws1_wb2+1):
        itemname_ws1_wb2 =ws1_wb2.cell(row=p, column=1).value
        if entry2 ==itemname_ws1_wb2:
            ivtqtyxl=ws1_wb2.cell(row=p, column=2).value
            ws1_wb2.cell(row=p, column=2).value = int(ivtqty) + int(ivtqtyxl)
            if int(price)!=0:
                ws1_wb2.cell(row=p, column=3).value =int(price)
                dic_prod[entry2]=int(price)
                lst_key_prod=list(dic_prod.keys())
            elif int(price)<=0:
                pass
            return True

def ivtaddbtn():
    global lst_key_prod
    global dic_prod
    entry2 = ivt_item_combo.get()
    ivtqty = ivt_item_qty.get()
    price=ivt_item_price.get()
    row_max_ws1_wb2 = int(ws1_wb2.max_row)
    if (updateinv()!=True):
        ws1_wb2.cell(row=row_max_ws1_wb2+1, column=1).value=str(entry2)
        ws1_wb2.cell(row=row_max_ws1_wb2+1, column=2).value=int(ivtqty)
        ws1_wb2.cell(row=row_max_ws1_wb2+1, column=3).value=int(price)
        dic_prod[entry2]=int(price)
        lst_key_prod=list(dic_prod.keys())
        ivt_item_combo.configure(values=lst_key_prod)
        item_combo.configure(values=lst_key_prod)
    wb2.save("inventory.xlsx")
    ivtrefreshbtn()


def submit():
    sum1()
    global xlrow
    if sheet_exists(c_d) == True:
        ws1_wb1 = wb1[str(c_d)]
        xlrprc = total_label.cget("text")
        new_data = [xlrprc]
        ws1_wb1.append(new_data)
        wb1.save("data.xlsx")
    else:
        ws1_wb1 = wb1.create_sheet(str(c_d))
        wb1.save("data.xlsx")
        xlrprc = total_label.cget("text")
        new_data = [xlrprc]
        ws1_wb1.append(new_data)
        wb1.save("data.xlsx")
    IvtReduc()
    clc1()

def day_book():
    ws1_wb1 = wb1[str(c_d)]
    daybook = 0
    rnm = ws1_wb1.max_row
    for i in range(1, rnm + 1):
        ab = ws1_wb1.cell(row=i, column=1).value
        daybook = daybook + int(ab)
    total_label.config(text=daybook)

def remove1():
    x = data_tree.selection()[0]
    data_tree.delete(x)

def removeinv():
    global dic_prod
    global lst_key_prod
    x = ivt_data_tree.selection()[0]
    rmlst=(ivt_data_tree.item(x)['values'])
    row_max_ws1_wb2 = int(ws1_wb2.max_row)
    for p in range(3, row_max_ws1_wb2+1):
        itemname_ws1_wb2 =ws1_wb2.cell(row=p, column=1).value
        if rmlst[0] ==itemname_ws1_wb2:
            ws1_wb2.delete_rows(p)
            del dic_prod[rmlst[0]]
            lst_key_prod.remove(rmlst[0])
            ivt_item_combo.configure(values=lst_key_prod)
            item_combo.configure(values=lst_key_prod)
            break
    wb2.save("inventory.xlsx")
    ivt_data_tree.delete(x)

def addbtn1():
    global count
    tempvar=None
    row_max_ws1_wb2 = ws1_wb2.max_row
    for p in range(3, row_max_ws1_wb2+1):
        itemname_ws1_wb2 =ws1_wb2.cell(row=p, column=1).value
        if(itemname_ws1_wb2==item_combo.get()):
            tempvar=int(p)
            break
    if ((tempvar!=None) and (int(item_qty.get())!=0) and (item_price.get()!='')):
        if int(item_qty.get())>int(ws1_wb2.cell(row=p, column=2).value):
            popup(0)
        else:
            data_tree.insert(parent="",index="end",iid=count,text="",values=(item_combo.get(),
            item_price.get(), item_qty.get()))
            count+=1
            item_combo.delete(0, END)
            item_price.delete(0, END)
            item_qty.delete(0, END)
            item_qty.insert(0, 0)
    else:
        popup(1)


# plus and minus button code
def plsbtn1(a):
    if a == 0:
        qty = int(item_qty.get())
        qty += 1000
        item_qty.delete(0, END)
        item_qty.insert(0, qty)
    elif a == 1:
        qty1 = int(ivt_item_qty.get())
        qty1 += 1000
        ivt_item_qty.delete(0, END)
        ivt_item_qty.insert(0, qty1)

def mnsbtn1(a):
    qty = int(item_qty.get())
    qty1 = int(ivt_item_qty.get())
    if a == 0:
        if qty > 1000:
            qty -= 1000
            item_qty.delete(0, END)
            item_qty.insert(0, qty)
        elif qty == 0:
            pass
    elif a == 1:
        if qty1 > 1000:
            qty1 -= 1000
            ivt_item_qty.delete(0, END)
            ivt_item_qty.insert(0, qty1)
        elif qty1 == 0:
            pass
    else:
        pass

def autoprice():
    entry1 = item_combo.get()
    if entry1!='':
        item_price.delete(0,END)
        item_price.insert(0,int(dic_prod[entry1]))
    else:
        pass

def IvtReduc():
    for line in data_tree.get_children():
        tempstore=data_tree.item(line)["values"]
        row_max_ws1_wb2 = int(ws1_wb2.max_row)
        for p in range(3, row_max_ws1_wb2+1):
            itemname_ws1_wb2 =ws1_wb2.cell(row=p, column=1).value
            if tempstore[0] == itemname_ws1_wb2:
                ivtqtyxl = ws1_wb2.cell(row=p, column=2).value
                ws1_wb2.cell(row=p, column=2).value = int(ivtqtyxl) - int(tempstore[2])
                break
            else:
                continue
    wb2.save("inventory.xlsx")

def sum1():
    global item_qty_l
    item_qty_l = []
    global item_price_l
    item_price_l = []

    for line in data_tree.get_children():
        item_qty_l.append((data_tree.item(line)["values"][2]))
        item_price_l.append((data_tree.item(line)["values"][1]))
    global s
    s = 0
    for i in range(len(item_price_l)):
        s = s + int(item_price_l[i]) * (int(item_qty_l[i]) / 1000)
    s = math.ceil(s)
    total_label.config(text=s)




# Creating window
window = Tk()
window.title("Retail-billing")

# main frame
frame = Frame(window)
frame.pack()

# first frame
first_frame = Frame(frame)
first_frame.grid(row=0, column=0, padx=10, pady=10)

homebtn = Button(first_frame, text="HOME", command=homebtn)
homebtn.grid(row=0, column=0, padx=13, pady=10)

invtrybtn = Button(first_frame, text="INVENTORY", command=inventorybtn)
invtrybtn.grid(row=1, column=0, padx=13, pady=10)

# createing notebook
notebook = ttk.Notebook(frame)
notebook.grid(row=0, column=1, padx=10)

home = Frame(notebook)
inventory = Frame(notebook)

updatedict()

notebook.add(home, text="HOME")
notebook.add(inventory, text="INVENTORY")
notebook.hide(1)

# inventory-frames
# frame1-inventory
ivt_data_frame = LabelFrame(inventory)
ivt_data_frame.grid(row=0, column=0, padx=10, pady=10)

ivt_item_label1 = Label(ivt_data_frame, text="Item's name")
ivt_item_label1.grid(row=0, column=0)

ivt_item_label2 = Label(ivt_data_frame, text="Quantity in grams")
ivt_item_label2.grid(row=0, column=1, columnspan=3)

ivt_item_label2 = Label(ivt_data_frame, text="Price")
ivt_item_label2.grid(row=0, column=4)

ivt_item_combo = ttk.Combobox(ivt_data_frame, values=lst_key_prod)
ivt_item_combo.grid(row=1, column=0, padx=13, pady=10)

ivt_item_qtybttn = Button(ivt_data_frame, text="-", command=lambda: mnsbtn1(1))
ivt_item_qtybttn.grid(row=1, column=1, padx=13, pady=10)

ivt_item_qty = Entry(ivt_data_frame)
ivt_item_qty.insert(0, 0)
ivt_item_qty.grid(row=1, column=2, padx=13, pady=10)

ivt_item_qtybttn = Button(ivt_data_frame, text="+", command=lambda: plsbtn1(1))
ivt_item_qtybttn.grid(row=1, column=3, padx=13, pady=10)

ivt_item_price = Entry(ivt_data_frame)
ivt_item_price.insert(0, 0)
ivt_item_price.grid(row=1, column=4, padx=13, pady=10)

ivt_add_btn = Button(
    ivt_data_frame, text="ADD", bg="cyan", fg="black", width=20, command=ivtaddbtn)
ivt_add_btn.grid(row=2, column=4, padx=13, pady=10)

ivt_rmv_btn = Button(ivt_data_frame,text="Remove",bg="cyan",fg="black",
width=20,command=removeinv)
ivt_rmv_btn.grid(row=2, column=0, padx=13, pady=10)

ivt_clr_btn = Button(ivt_data_frame,text="Reset",bg="cyan",fg="black",
width=20,command=reset)
ivt_clr_btn.grid(row=2, column=1, padx=13, pady=10,columnspan=3)

# frame2-inventory
ivt_detail_frame = LabelFrame(inventory, text="INVENTORY")
ivt_detail_frame.grid(row=2, column=0)

# scrollbar code
ivt_tree_scroll = Scrollbar(ivt_detail_frame)
ivt_tree_scroll.pack(side=RIGHT, fill=Y)

# treeview
ivt_data_tree = ttk.Treeview(
    ivt_detail_frame, yscrollcommand=ivt_tree_scroll.set, selectmode="extended")
ivt_data_tree["columns"] = ("Name", "Quantity","Price")
ivt_data_tree.column("#0", width=0)
ivt_data_tree.column("Name", anchor=W, width=185)
ivt_data_tree.column("Quantity", anchor=W, width=185)
ivt_data_tree.column("Price", anchor=W, width=185)

ivt_data_tree.heading("#0", text="", anchor=W)
ivt_data_tree.heading("Name", text="Item", anchor=W)
ivt_data_tree.heading("Quantity", text="Qty", anchor=W)
ivt_data_tree.heading("Price", text="Price", anchor=W)
ivt_data_tree.pack(pady=20, padx=20)

# config scroll bar
ivt_tree_scroll.config(command=ivt_data_tree.yview)


# frame1-home
total_frame = LabelFrame(home, text=("Calculations made easy"))
total_frame.grid(row=0, column=0, padx=10, pady=10)

total_label = Label(total_frame,text="Please Add Items",bg="black",fg="white",
    font=("Arial", 20),width=35)
total_label.grid(row=0, column=0)

for widget in total_frame.winfo_children():
    widget.grid_configure(padx=2, pady=2)

# frame2-home
# FINAL SUM LABEL
data_frame = LabelFrame(home, text="Add items and click SUM")
data_frame.grid(row=1, column=0, padx=10, pady=10)

# row1 -labels
item_label1 = Label(data_frame, text="Item's name")
item_label1.grid(row=0, column=0)
item_label1 = Label(data_frame, text="Price")
item_label1.grid(row=0, column=1)
item_label1 = Label(data_frame, text="Quantity in grams")
item_label1.grid(row=0, column=2, columnspan=3)

# SELECTION
# row2-combobox,entry,buttons
item_combo = ttk.Combobox(data_frame, values=lst_key_prod)
item_combo.grid(row=1, column=0, padx=13, pady=10)

item_price = Entry(data_frame)
item_price.insert(0,0)
item_price.grid(row=1, column=1, padx=13, pady=10)

item_qtybttn = Button(data_frame, text="-", command=lambda: mnsbtn1(0))
item_qtybttn.grid(row=1, column=2, padx=13, pady=10)

item_qty = Entry(data_frame)
item_qty.insert(0, 0)
item_qty.grid(row=1, column=3, padx=13, pady=10)

item_qtybttn = Button(data_frame, text="+", command=lambda: plsbtn1(0))
item_qtybttn.grid(row=1, column=4, padx=13, pady=10)

add_btn = Button(data_frame, text="ADD", bg="cyan", fg="black", width=20, command=addbtn1)
add_btn.grid(row=2, column=2, padx=13, pady=10, columnspan=3)

dn_btn = Button(data_frame, text="AUTO PRICE", bg="yellow", fg="black", width=20, command=autoprice)
dn_btn.grid(row=2, column=1, padx=13, pady=10, columnspan=2)

rmv_btn = Button(data_frame, text="REMOVE", bg="red", fg="black", width=20, command=remove1)
rmv_btn.grid(row=2, column=0, padx=13, pady=10)

# frame3-home
detail_frame = LabelFrame(home)
detail_frame.grid(row=3, column=0,padx=13, pady=10)

# frame4-home(end frame)
end_frame = LabelFrame(home)
end_frame.grid(row=4, column=0,padx=13, pady=10)

sum_btn = Button(end_frame, text="SUM", bg="green2", fg="black", width=15, command=sum1)
sum_btn.grid(row=0, column=1)

clc_btn = Button(end_frame, text="CLEAR", bg="red", fg="black", width=15, command=clc1)
clc_btn.grid(row=0, column=0, sticky="W")

subm_btn = Button(end_frame, text="SUBMIT", bg="yellow", fg="black", width=15, command=submit)
subm_btn.grid(row=0, column=3)

load_btn = Button(end_frame, text="DAY BOOK", bg="SeaGreen1", fg="black", width=15, command=day_book)
load_btn.grid(row=0, column=2, sticky="E")

for widget in end_frame.winfo_children():
    widget.grid_configure(padx=13, pady=10)

# scrollbar code
tree_scroll = Scrollbar(detail_frame)
tree_scroll.pack(side=RIGHT, fill=Y)

# treeview-frame3-home
data_tree = ttk.Treeview(detail_frame, yscrollcommand=tree_scroll.set, selectmode="extended")
data_tree["columns"] = ("Name", "Rate", "Quantity")
data_tree.column("#0", width=0)
data_tree.column("Name", anchor=W, width=165)
data_tree.column("Rate", anchor=W, width=165)
data_tree.column("Quantity", anchor=W, width=165)

data_tree.heading("#0", text="", anchor=W)
data_tree.heading("Name", text="Item", anchor=W)
data_tree.heading("Rate", text="Rate", anchor=W)
data_tree.heading("Quantity", text="Qty", anchor=W)
data_tree.pack(pady=20, padx=20)

# config scroll bar
tree_scroll.config(command=data_tree.yview)

window.mainloop()

